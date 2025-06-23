import io
import csv
import logging
from typing import List, Optional, BinaryIO
from datetime import date, datetime
from decimal import Decimal
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from sqlalchemy.ext.asyncio import AsyncSession

from src.database.models import Transaction, User, Category
from src.services.transaction import TransactionService
from src.services.category import CategoryService
from src.services.s3_storage import S3StorageService
from src.utils.text_parser import ExpenseParser

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting transaction data"""
    
    def __init__(self):
        self.transaction_service = TransactionService()
        self.category_service = CategoryService()
        self.s3_service = S3StorageService()
        self.expense_parser = ExpenseParser()
        
        # Try to register fonts for PDF with Cyrillic support
        self._register_fonts()
    
    def _register_fonts(self):
        """Register fonts with Cyrillic support for PDF generation"""
        import os
        from pathlib import Path
        
        # First, try to use bundled font
        project_root = Path(__file__).parent.parent.parent
        bundled_font_path = project_root / 'assets' / 'fonts' / 'NotoSans-Regular.ttf'
        
        if bundled_font_path.exists():
            try:
                pdfmetrics.registerFont(TTFont('NotoSans', str(bundled_font_path)))
                self.pdf_font = 'NotoSans'
                logger.info(f"Successfully registered bundled font: {bundled_font_path}")
                return
            except Exception as e:
                logger.warning(f"Failed to register bundled font: {e}")
        
        # If bundled font not found, try system fonts
        import platform
        
        font_paths = []
        
        if platform.system() == 'Darwin':  # macOS
            font_paths = [
                '/System/Library/Fonts/Helvetica.ttc',
                '/System/Library/Fonts/Arial Unicode.ttf',
                '/Library/Fonts/Arial Unicode.ttf',
                '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
                # DejaVu fonts if installed via homebrew
                '/opt/homebrew/share/fonts/DejaVuSans.ttf',
                '/usr/local/share/fonts/DejaVuSans.ttf',
            ]
        elif platform.system() == 'Linux':
            font_paths = [
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
            ]
        
        # Try to register a system font
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('CustomFont', font_path))
                    self.pdf_font = 'CustomFont'
                    logger.info(f"Successfully registered system font: {font_path}")
                    return
                except Exception as e:
                    logger.debug(f"Failed to register font {font_path}: {e}")
                    continue
        
        # Last resort - download font
        try:
            import urllib.request
            
            font_url = "https://github.com/googlefonts/noto-fonts/raw/main/hinted/ttf/NotoSans/NotoSans-Regular.ttf"
            
            # Ensure fonts directory exists
            fonts_dir = project_root / 'assets' / 'fonts'
            fonts_dir.mkdir(parents=True, exist_ok=True)
            
            # Download to permanent location
            download_path = fonts_dir / 'NotoSans-Regular.ttf'
            
            if not download_path.exists():
                logger.info("Downloading Noto Sans font for Cyrillic support...")
                urllib.request.urlretrieve(font_url, str(download_path))
            
            pdfmetrics.registerFont(TTFont('NotoSans', str(download_path)))
            self.pdf_font = 'NotoSans'
            logger.info("Successfully downloaded and registered Noto Sans font")
        except Exception as e:
            logger.warning(f"Failed to download font: {e}")
            # Use Helvetica as last resort
            self.pdf_font = 'Helvetica'
            logger.warning("Using Helvetica font - Cyrillic text may not display correctly")
    
    async def export_transactions(
        self,
        session: AsyncSession,
        user: User,
        format: str,
        start_date: date,
        end_date: date,
        category_ids: Optional[List[str]] = None
    ) -> Optional[str]:
        """Export transactions in specified format and upload to S3"""
        # Get transactions
        transactions = await self._get_transactions_for_export(
            session, user.id, start_date, end_date, category_ids
        )
        
        if not transactions:
            return None
        
        # Get categories for mapping
        categories = await self.category_service.get_user_categories(session, user.id, active_only=False)
        category_map = {cat.id: cat for cat in categories}
        
        # Export based on format
        file_io = None
        filename = None
        content_type = None
        
        if format == 'xlsx':
            file_io = await self._export_to_excel(transactions, category_map, user)
            filename = f"expenses_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif format == 'csv':
            file_io = await self._export_to_csv(transactions, category_map, user)
            filename = f"expenses_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.csv"
            content_type = 'text/csv'
        elif format == 'pdf':
            file_io = await self._export_to_pdf(transactions, category_map, user, start_date, end_date)
            filename = f"expenses_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.pdf"
            content_type = 'application/pdf'
        else:
            raise ValueError(f"Unsupported export format: {format}")
        
        if not file_io:
            return None
        
        # Upload to S3 if enabled
        if self.s3_service.enabled:
            file_io.seek(0)
            s3_url = await self.s3_service.upload_export_file(
                user_id=user.id,
                file_data=file_io.getvalue(),
                filename=filename,
                content_type=content_type
            )
            
            if s3_url:
                logger.info(f"[S3] Export file uploaded: {s3_url}")
                return s3_url
        
        # Fallback: return file data as string (for backward compatibility)
        file_io.seek(0)
        return file_io.getvalue()
    
    async def _get_transactions_for_export(
        self,
        session: AsyncSession,
        user_id: int,
        start_date: date,
        end_date: date,
        category_ids: Optional[List[str]] = None
    ) -> List[Transaction]:
        """Get transactions for export with filtering"""
        transactions = []
        
        # If specific categories requested
        if category_ids:
            for category_id in category_ids:
                cat_transactions = await self.transaction_service.get_user_transactions(
                    session, user_id,
                    start_date=start_date,
                    end_date=end_date,
                    category_id=category_id,
                    limit=10000  # High limit for export
                )
                transactions.extend(cat_transactions)
        else:
            # Get all transactions
            transactions = await self.transaction_service.get_user_transactions(
                session, user_id,
                start_date=start_date,
                end_date=end_date,
                limit=10000
            )
        
        # Sort by date
        transactions.sort(key=lambda x: x.transaction_date, reverse=True)
        
        return transactions
    
    async def _export_to_excel(
        self,
        transactions: List[Transaction],
        category_map: dict,
        user: User
    ) -> BinaryIO:
        """Export transactions to Excel format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Транзакции"
        
        # Headers
        headers = [
            "Дата", "Время", "Категория", "Описание", "Место",
            "Сумма", "Валюта", "Сумма в основной валюте", "Курс"
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, tx in enumerate(transactions, 2):
            category = category_map.get(tx.category_id)
            category_name = f"{category.icon} {category.get_name(user.language_code)}" if category else "?"
            
            ws.cell(row=row, column=1, value=tx.transaction_date.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=2, value=tx.transaction_date.strftime('%H:%M'))
            ws.cell(row=row, column=3, value=category_name)
            ws.cell(row=row, column=4, value=tx.description or "")
            ws.cell(row=row, column=5, value=tx.merchant or "")
            ws.cell(row=row, column=6, value=float(tx.amount))
            ws.cell(row=row, column=7, value=tx.currency)
            ws.cell(row=row, column=8, value=float(tx.amount_primary))
            ws.cell(row=row, column=9, value=float(tx.exchange_rate))
        
        # Add summary
        summary_row = len(transactions) + 3
        ws.cell(row=summary_row, column=1, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=summary_row, column=8, value=f"=SUM(H2:H{len(transactions)+1})").font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    async def _export_to_csv(
        self,
        transactions: List[Transaction],
        category_map: dict,
        user: User
    ) -> BinaryIO:
        """Export transactions to CSV format"""
        output = io.StringIO()
        
        fieldnames = [
            'date', 'time', 'category', 'description', 'merchant',
            'amount', 'currency', 'amount_primary', 'exchange_rate'
        ]
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for tx in transactions:
            category = category_map.get(tx.category_id)
            category_name = f"{category.icon} {category.get_name(user.language_code)}" if category else "?"
            
            writer.writerow({
                'date': tx.transaction_date.strftime('%d.%m.%Y'),
                'time': tx.transaction_date.strftime('%H:%M'),
                'category': category_name,
                'description': tx.description or '',
                'merchant': tx.merchant or '',
                'amount': str(tx.amount),
                'currency': tx.currency,
                'amount_primary': str(tx.amount_primary),
                'exchange_rate': str(tx.exchange_rate)
            })
        
        # Convert to bytes
        output_bytes = io.BytesIO()
        output_bytes.write(output.getvalue().encode('utf-8-sig'))  # BOM for Excel
        output_bytes.seek(0)
        
        return output_bytes
    
    async def _export_to_pdf(
        self,
        transactions: List[Transaction],
        category_map: dict,
        user: User,
        start_date: date,
        end_date: date
    ) -> BinaryIO:
        """Export transactions to PDF format"""
        output = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Update font for all styles to support Cyrillic
        for style_name in styles.byName:
            styles[style_name].fontName = self.pdf_font
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=self.pdf_font,
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Title
        title = Paragraph(
            f"Отчет о расходах<br/>{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
            title_style
        )
        story.append(title)
        story.append(Spacer(1, 0.5 * inch))
        
        # Summary statistics
        total_amount = sum(tx.amount_primary for tx in transactions)
        category_totals = {}
        
        for tx in transactions:
            cat_id = tx.category_id
            if cat_id not in category_totals:
                category_totals[cat_id] = Decimal('0')
            category_totals[cat_id] += tx.amount_primary
        
        # Summary table
        summary_data = [['Показатель', 'Значение']]
        summary_data.append(['Всего транзакций', str(len(transactions))])
        summary_data.append(['Общая сумма', self.expense_parser.format_amount(total_amount, user.primary_currency)])
        
        if transactions:
            avg_amount = total_amount / len(transactions)
            summary_data.append(['Средний чек', self.expense_parser.format_amount(avg_amount, user.primary_currency)])
        
        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), self.pdf_font),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 0.5 * inch))
        
        # Category breakdown
        if category_totals:
            story.append(Paragraph("Расходы по категориям", styles['Heading2']))
            
            cat_data = [['Категория', 'Сумма', 'Процент']]
            
            for cat_id, amount in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
                category = category_map.get(cat_id)
                if category:
                    cat_name = f"{category.icon} {category.get_name(user.language_code)}"
                else:
                    cat_name = "?"
                
                percentage = (amount / total_amount * 100) if total_amount > 0 else 0
                cat_data.append([
                    cat_name,
                    self.expense_parser.format_amount(amount, user.primary_currency),
                    f"{percentage:.1f}%"
                ])
            
            cat_table = Table(cat_data, colWidths=[2.5 * inch, 2 * inch, 1 * inch])
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), self.pdf_font),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(cat_table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        return output